import streamlit as st
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.cell.cell import MergedCell # 判定用にインポート
import json
import io
import os
import re

# ページ設定
st.set_page_config(page_title="指導案作成WEBアプリ", layout="wide")

# ==========================================
# 0. ユーティリティ関数（エラー回避用）
# ==========================================
def safe_write(ws, cell_address, value):
    """
    結合セルエラー（MergedCell...read-only）を回避して書き込む関数。
    指定したセルが結合の一部（左上以外）だった場合、自動的に左上のセルを探して書き込む。
    """
    try:
        # まず普通に書き込みを試みる（対象がセルオブジェクトの場合）
        if isinstance(ws[cell_address], MergedCell):
            # 対象が結合セル(MergedCell)の場合、ここには書き込めない
            # そのセルが含まれる「結合範囲」を探す
            for merged_range in ws.merged_cells.ranges:
                if cell_address in merged_range:
                    # 結合範囲の左上（start_cell）を取得
                    top_left_coord = merged_range.start_cell.coordinate
                    # 左上のセルに値を書き込む
                    ws[top_left_coord] = value
                    # 書式設定（左上揃え・折り返し）
                    ws[top_left_coord].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                    return
        else:
            # 結合セルでない、または結合の左上セルの場合は普通に書き込む
            ws[cell_address] = value
            ws[cell_address].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

    except Exception as e:
        st.warning(f"セル {cell_address} への書き込み中に警告: {e}")

# ==========================================
# 1. プロンプト生成ロジック
# ==========================================
def generate_prompt_text(data):
    prompt = f"""
あなたは特別支援学校および公立学校における熟練の教員です。
以下の【授業情報】を基に、学習指導案に必要な情報を補完し、指定の【JSON形式】のみで出力してください。
余計な挨拶や解説は不要です。JSONデータだけを返してください。

■ 【授業情報】
[必須項目]
・学部学年: {data['grade']}
・教科単元: {data['subject']}
・日時: {data['date']}
・時間: {data['time']}
・場所: {data['place']}
・本時の内容: {data['content']}

[任意項目]
・目標: {data['goals_in'] if data['goals_in'] else "未定（文脈に合わせて最大3つ生成せよ）"}
・評価の基準: {data['eval_in'] if data['eval_in'] else "未定（3観点を含めて生成せよ）"}
・学習内容のヒント: {data['flow_in'] if data['flow_in'] else "未定（自然な流れで構成せよ）"}

■ 【出力フォーマット（厳守）】
以下のJSON構造を崩さずに返してください。
{{
  "basic_info": {{
    "grade": "{data['grade']}",
    "subject": "{data['subject']}",
    "date": "{data['date']}",
    "time": "{data['time']}",
    "place": "{data['place']}",
    "content": "{data['content']}"
  }},
  "goals": ["目標1", "目標2", "目標3"],
  "evaluation": ["評価基準1", "評価基準2", "評価基準3"],
  "flow": [
    {{
      "time": "5",
      "activity": "導入：挨拶...",
      "notes": "留意点..."
    }},
    {{
      "time": "10",
      "activity": "展開1：...",
      "notes": "..."
    }}
  ],
  "materials": "準備物リスト"
}}
"""
    return prompt

# ==========================================
# 2. Excel出力ロジック
# ==========================================
def create_excel(template_path, json_data):
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
    except Exception as e:
        return None, f"テンプレート読み込みエラー: {e}"

    # --- ① 基本情報の入力（safe_writeを使用） ---
    bi = json_data.get('basic_info', {})
    
    safe_write(ws, 'C2', bi.get('grade', ''))      # 学部学年
    safe_write(ws, 'I2', bi.get('subject', ''))    # 教科単元
    safe_write(ws, 'C3', bi.get('date', ''))       # 日時
    safe_write(ws, 'K3', bi.get('time', ''))       # 時間
    safe_write(ws, 'N3', bi.get('place', ''))      # 場所
    safe_write(ws, 'C4', bi.get('content', ''))    # 本時の内容

    # --- ② 目標（B10, B11, B12） ---
    goals = json_data.get('goals', [])
    if len(goals) > 0: safe_write(ws, 'B10', f"・{goals[0]}")
    if len(goals) > 1: safe_write(ws, 'B11', f"・{goals[1]}")
    if len(goals) > 2: safe_write(ws, 'B12', f"・{goals[2]}")

    # --- ③ 評価の基準（B14） ---
    evals = json_data.get('evaluation', [])
    eval_text = "\n".join([f"・{e}" for e in evals])
    safe_write(ws, 'B14', eval_text)

    # --- ④ 本時の展開（A13～ 1行あけ） ---
    flow_list = json_data.get('flow', [])
    current_row = 13
    
    for item in flow_list:
        # 時間 (A列)
        safe_write(ws, f'A{current_row}', item.get('time', ''))

        # 学習内容 (B列:J列想定)
        # ※ここがエラーの原因になりやすい場所です。
        # テンプレートでB13:J13が結合されているなら 'B13' に書き込めばOK。
        # もし 'C13' などが指定されるとエラーになりますが、safe_writeが救ってくれます。
        safe_write(ws, f'B{current_row}', item.get('activity', ''))

        # 留意点 (K列:M列想定)
        safe_write(ws, f'K{current_row}', item.get('notes', ''))

        # 次の項目は1行空ける
        current_row += 2 

    # --- 準備物 (N13) ---
    safe_write(ws, 'N13', json_data.get('materials', ''))

    # 保存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, None

# ==========================================
# 3. メイン画面 UI
# ==========================================
st.title("📝 指導案作成WEBアプリ")
st.markdown("ChatGPTやGeminiを使って指導案を作成し、Excelに出力します。")

# --- Step 1: 情報入力 ---
st.header("1. 基本情報を入力")
with st.container():
    col1, col2 = st.columns(2)
    with col1:
        in_grade = st.text_input("学部学年", "小学部 5年")
        in_subject = st.text_input("教科単元", "生活単元学習「お祭りを開こう」")
        in_date = st.text_input("日時", "令和6年11月20日")
    with col2:
        in_time = st.text_input("時間", "45分")
        in_place = st.text_input("場所", "5年1組教室")
        in_content = st.text_input("本時の内容", "模擬店の商品作り")

    with st.expander("詳細設定（任意入力）- 空欄でもAIが補完します"):
        in_goals = st.text_area("目標（最大3つ）", height=68)
        in_eval = st.text_area("評価の基準", height=68)
        in_flow = st.text_area("学習内容・メモ", height=100)

# データをまとめる
input_data = {
    "grade": in_grade, "subject": in_subject, "date": in_date,
    "time": in_time, "place": in_place, "content": in_content,
    "goals_in": in_goals, "eval_in": in_eval, "flow_in": in_flow
}

# --- Step 2: プロンプト生成 ---
st.header("2. AI用プロンプトを生成")
if st.button("プロンプト作成 📋"):
    prompt_text = generate_prompt_text(input_data)
    st.code(prompt_text, language="text")
    st.success("コピーしてChatGPTやGeminiに貼り付けてください。")

# --- Step 3: AI出力の貼り付け ---
st.header("3. AIからの回答を貼り付け")
json_input_str = st.text_area("ここにAIの回答をペースト", height=300)

# --- Step 4: Excel生成 ---
st.header("4. 指導案Excelのダウンロード")

if st.button("Excel作成実行 🚀"):
    if not json_input_str.strip():
        st.error("AIの回答が貼り付けられていません。")
    else:
        try:
            # 1. JSONのクリーニングと解析
            clean_json = re.sub(r"```json\s*|\s*```", "", json_input_str).strip()
            start_idx = clean_json.find('{')
            end_idx = clean_json.rfind('}') + 1
            if start_idx != -1 and end_idx != -1:
                clean_json = clean_json[start_idx:end_idx]
            
            data_dict = json.loads(clean_json)
            
            # 2. ファイルパスの自動解決（pages対策）
            # このファイル(pages/app.py)のあるフォルダを取得
            current_dir = os.path.dirname(os.path.abspath(__file__))
            # 一つ上の階層(ルート)を取得
            base_dir = os.path.dirname(current_dir)
            # ルートにあるExcelファイルを指定
            template_file = os.path.join(base_dir, "指導案.xlsx")
            
            # デバッグ用：パスが見つかるか確認（見つからなければエラー表示）
            if not os.path.exists(template_file):
                # もしルートになければ、同じフォルダ(pages)を探す予備処理
                template_file = os.path.join(current_dir, "指導案.xlsx")

            if not os.path.exists(template_file):
                st.error(f"エラー: テンプレートファイルが見つかりません。\n探した場所:\n1. {os.path.join(base_dir, '指導案.xlsx')}\n2. {os.path.join(current_dir, '指導案.xlsx')}")
            else:
                # 3. Excel生成実行
                excel_data, err = create_excel(template_file, data_dict)
                if err:
                    st.error(err)
                else:
                    st.success("成功！ダウンロードできます。")
                    st.download_button(
                        label="📥 指導案.xlsx をダウンロード",
                        data=excel_data,
                        file_name="完成_指導案.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
        except json.JSONDecodeError:
            st.error("JSON解析エラー: AIの回答を正しく貼り付けてください。")
        except Exception as e:
            st.error(f"予期せぬエラー: {e}")