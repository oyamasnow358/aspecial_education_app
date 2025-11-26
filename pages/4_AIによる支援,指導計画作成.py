import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.cell.cell import MergedCell
import json
import io
import os
import re
import base64
from pathlib import Path
from io import BytesIO

# ==========================================
# 0. ページ設定
# ==========================================

st.set_page_config(
    page_title="AIエージェントによる個別の支援計画・指導計画作成サポート",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ==========================================
# 1. ユーティリティ関数（画像・Excel）
# ==========================================

# --- 画像処理 (ロゴ読み込み) ---
def get_img_as_base64(file):
    try:
        # 画像パスを絶対パスで解決
        script_path = Path(__file__)
        # 修正: カレントディレクトリ(parent)と親ディレクトリ(parent.parent)の両方を探すように変更
        possible_paths = [script_path.parent / file, script_path.parent.parent / file]
        
        for img_path in possible_paths:
            if img_path.exists():
                with open(img_path, "rb") as f:
                    data = f.read()
                return base64.b64encode(data).decode()
        return None
    except:
        return None

# ロゴファイル設定 (あれば表示、なければプレースホルダー)
logo_path = "mirairo2.png" 
logo_b64 = get_img_as_base64(logo_path)
logo_html = f'<img src="data:image/png;base64,{logo_b64}" class="logo-img">' if logo_b64 else '<div class="logo-placeholder">🤖</div>'

# --- Excel書き込み用 ---
def safe_write(ws, cell_address, value):
    """
    結合セルエラー（MergedCell...read-only）を回避して書き込む関数。
    書き込み時に「折り返し全体を表示」「上揃え」「左揃え」を適用します。
    """
    try:
        if value is None:
            value = ""
        value = str(value)

        if isinstance(ws[cell_address], MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell_address in merged_range:
                    top_left_coord = merged_range.start_cell.coordinate
                    ws[top_left_coord] = value
                    ws[top_left_coord].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                    return
        else:
            ws[cell_address] = value
            ws[cell_address].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    except Exception as e:
        st.warning(f"⚠️ セル {cell_address} への書き込み中に警告: {e}")

# ==========================================
# 2. デザイン定義 (Mirairoスタイル)
# ==========================================
def load_css():
    st.markdown("""
        <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@400;700;900&display=swap" rel="stylesheet">
    """, unsafe_allow_html=True)
    
    css = f"""
    <style>
        /* --- 全体フォント --- */
        html, body, [class*="css"] {{
            font-family: 'Noto Sans JP', sans-serif !important;
            color: #333333 !important;
        }}

        /* --- 背景 (白92%透過・画像あり) --- */
        [data-testid="stAppViewContainer"] {{
            background-color: #ffffff;
            background-image: linear-gradient(rgba(255,255,255,0.92), rgba(255,255,255,0.92)), url("https://i.imgur.com/AbUxfxP.png");
            background-size: cover;
            background-attachment: fixed;
        }}

        /* --- 文字色 (濃紺・くっきり) --- */
        h1, h2, h3, h4, h5, h6 {{
            color: #0f172a !important; /* 濃いネイビーブラック */
            text-shadow: none !important;
        }}
        p, span, div, label {{
            color: #333333 !important;
            text-shadow: none !important;
        }}

        /* --- サイドバー (すりガラス効果) --- */
        [data-testid="stSidebar"] {{
            background-color: rgba(255, 255, 255, 0.85) !important;
            backdrop-filter: blur(20px);
            -webkit-backdrop-filter: blur(20px);
            border-right: 1px solid #e2e8f0 !important;
        }}
        [data-testid="stSidebar"] * {{
            color: #333333 !important;
        }}

        /* 
           ================================================================
           ★ アニメーション定義 (下からフワッと)
           ================================================================
        */
        @keyframes fadeInUp {{
            from {{ opacity: 0; transform: translateY(40px); }}
            to {{ opacity: 1; transform: translateY(0); }}
        }}

        /* 
           ================================================================
           ★ コンテナデザイン (白背景・影付き・アニメーション)
           ================================================================
           st.container(border=True) のスタイルをオーバーライド
        */
        [data-testid="stBorderContainer"] {{
            background-color: #ffffff !important;
            border: 2px solid #e2e8f0 !important; /* 薄いグレーの枠線 */
            border-radius: 15px !important;
            padding: 25px !important;
            margin-bottom: 20px !important;
            box-shadow: 0 4px 12px rgba(0,0,0,0.05) !important;
            
            /* アニメーション適用 */
            opacity: 0; 
            animation-name: fadeInUp;
            animation-duration: 0.8s;
            animation-fill-mode: forwards;
            animation-timing-function: cubic-bezier(0.2, 0.8, 0.2, 1);
        }}
        
        /* コンテナの出現タイミングをずらす */
        div.element-container:nth-of-type(1) [data-testid="stBorderContainer"] {{ animation-delay: 0.1s; }}
        div.element-container:nth-of-type(2) [data-testid="stBorderContainer"] {{ animation-delay: 0.2s; }}
        div.element-container:nth-of-type(3) [data-testid="stBorderContainer"] {{ animation-delay: 0.3s; }}
        div.element-container:nth-of-type(4) [data-testid="stBorderContainer"] {{ animation-delay: 0.4s; }}
        div.element-container:nth-of-type(5) [data-testid="stBorderContainer"] {{ animation-delay: 0.5s; }}

        [data-testid="stBorderContainer"]:hover {{
            border-color: #4a90e2 !important;
            background-color: #f8fafc !important;
            transform: translateY(-2px);
            box-shadow: 0 10px 25px rgba(74, 144, 226, 0.15) !important;
            transition: all 0.3s ease;
        }}

        /* --- ボタン --- */
        .stButton > button {{
            width: 100%;
            background-color: #ffffff !important;
            border: 2px solid #e2e8f0 !important;
            color: #4a90e2 !important;
            font-weight: bold !important;
            border-radius: 30px !important; /* 丸みを強く */
            padding: 10px !important;
            transition: all 0.3s ease !important;
        }}
        .stButton > button:hover {{
            background-color: #4a90e2 !important;
            color: #ffffff !important;
            border-color: #4a90e2 !important;
            transform: scale(1.02);
        }}
        
        /* Primaryボタン (強調) */
        .stButton > button[kind="primary"] {{
            background-color: #4a90e2 !important;
            color: #ffffff !important;
            border: 2px solid #4a90e2 !important;
        }}
        .stButton > button[kind="primary"]:hover {{
            background-color: #2563eb !important;
            border-color: #2563eb !important;
        }}

        /* --- 入力フォーム --- */
        .stTextInput input, .stTextArea textarea, .stSelectbox div[data-baseweb="select"] {{
            background-color: #ffffff !important;
            color: #333333 !important;
            border: 1px solid #cbd5e1 !important;
            border-radius: 8px !important;
        }}
        .stTextInput input:focus, .stTextArea textarea:focus {{
            border-color: #4a90e2 !important;
            box-shadow: 0 0 0 2px rgba(74,144,226,0.2) !important;
        }}

        /* --- ラジオボタン（モード選択） --- */
        div[role="radiogroup"] {{
            background-color: #f8fafc;
            padding: 15px;
            border-radius: 12px;
            border: 2px solid #4a90e2;
            margin-bottom: 20px;
        }}

        /* --- エキスパンダー (アコーディオン) --- */
        .streamlit-expanderHeader {{
            background-color: #f1f5f9 !important;
            color: #0f172a !important;
            font-weight: bold !important;
            border-radius: 8px !important;
            border: 1px solid #e2e8f0;
        }}
        .streamlit-expanderContent {{
            border: none;
            color: #333 !important;
        }}

        /* --- 戻るボタン (指定デザイン) --- */
        .back-link {{
            margin-bottom: 20px;
        }}
        .back-link a {{
            display: inline-block;
            padding: 10px 20px;
            background: #ffffff;
            border: 2px solid #e2e8f0;
            border-radius: 25px;
            color: #4a90e2 !important;
            text-decoration: none;
            font-weight: bold;
            transition: all 0.3s;
            box-shadow: 0 2px 5px rgba(0,0,0,0.05);
        }}
        .back-link a:hover {{
            background: #4a90e2;
            color: #ffffff !important;
            border-color: #4a90e2;
            box-shadow: 0 4px 10px rgba(74, 144, 226, 0.2);
        }}

        /* --- ヘッダーレイアウト (修正: 中央寄せを追加) --- */
        .header-container {{
            display: flex;
            align-items: center;
            justify-content: center; /* ★ここを追加しました */
            gap: 20px;
            margin-bottom: 30px;
            padding-bottom: 20px;
            border-bottom: 2px solid #f1f5f9;
            animation: fadeInUp 0.8s ease-out forwards;
        }}
        .logo-img {{
            width: 70px;
            height: auto;
            filter: drop-shadow(0 4px 6px rgba(0,0,0,0.1));
        }}
        .logo-placeholder {{
            font-size: 3rem;
            margin-right: 15px;
        }}
        .page-title {{
            font-size: 2rem;
            font-weight: 900;
            color: #0f172a;
            margin: 0;
            line-height: 1.2;
        }}
        
        /* コードブロック */
        code {{
            background-color: #f1f5f9 !important;
            color: #0f172a !important;
            border: 1px solid #e2e8f0;
            border-radius: 8px;
        }}
    </style>
    """
    st.markdown(css, unsafe_allow_html=True)

load_css()

# ==========================================
# 3. メインコンテンツ
# ==========================================

# --- 戻るボタン (★指定のHTML) ---
st.markdown('<div class="back-link"><a href="https://aspecialeducationapp-6iuvpdfjbflp4wyvykmzey.streamlit.app/" target="_self">« TOPページに戻る</a></div>', unsafe_allow_html=True)

# --- ヘッダーエリア ---
st.markdown(f"""
    <div class="header-container">
        {logo_html}
        <div>
            <h1 class="page-title">個別の支援計画・指導計画作成サポート</h1>
            <p style="color:#475569; margin:0; font-weight:bold;">AIを活用して計画作成を効率化します</p>
        </div>
    </div>
""", unsafe_allow_html=True)

# ==========================================
# ★ モード選択エリア
# ==========================================
st.markdown("### 🛠️ 利用モードを選択してください")
mode_selection = st.radio(
    "モード選択",
    ("📋 通常モード (プロンプトをコピーして使う)", "🚀 Excel作成モード (岩槻はるかぜ機能)"),
    index=0,
    horizontal=True,
    label_visibility="collapsed"
)

# モード判定フラグ
is_excel_mode = "Excel" in mode_selection

# モード説明
if is_excel_mode:
    st.info("🚀 **Excel作成モード**が選択されています。\n入力内容は通常モードと全く同じですが、AIへの出力指定のみが**JSON形式**に変更されます。生成されたJSONコードをページ下部の入力欄に貼り付けることで、Excelを自動作成できます。")
else:
    st.success("📋 **通常モード**が選択されています。\nAIは読みやすい文章形式で回答します。Word等に手動でコピペする方に適しています。")

st.markdown("---")

# --- AIチャットへのリンク ---
with st.container(border=True):
    st.markdown("""
    <div style="text-align: center; margin-bottom: 10px;">
        <h3 style="margin: 0; padding: 0; color: #0f172a; border: none; font-size: 1.4em;">
            🚀 プロンプトをコピーしたら、AIチャットへ！
        </h3>
        <p style="margin-top: 5px; color: #555; font-size: 0.95em;">
            下のボタンを押すと各AIチャットが開きます。コピーしたプロンプトを貼り付けてください。
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # リンクボタンを横並び（2カラム）に配置
    btn_col1, btn_col2 = st.columns(2)
    with btn_col1:
        st.link_button("ChatGPT を開く ↗", "https://chat.openai.com/", type="primary", use_container_width=True)
    with btn_col2:
        st.link_button("Gemini を開く ↗", "https://gemini.google.com/", type="primary", use_container_width=True)

st.markdown("<br>", unsafe_allow_html=True)

# ==========================================
# プロンプト①
# ==========================================
with st.expander("プロンプト①【プランA：特別な教育的ニーズ／合理的配慮】"):
    with st.container(border=True):
        st.subheader("プロンプト①【プランA：特別な教育的ニーズ／合理的配慮】")
        jittai_1 = st.text_area("✅ お子さんの実態や課題を入力", value="視力が弱い、落ち着きがない、疲れやすい、音に敏感、話しかけられると混乱する、同じ行動を繰り返す", height=100, key="jittai_1")
        
        if st.button("プロンプト① を生成", key="btn_1", use_container_width=True):
            
            # ★ 共通の指示内容（ここを絶対に変更しない）
            common_instructions = """
【出力項目】：
1.特別な教育的ニーズ

・入力から共通性を見出し、2～3つの抽象的カテゴリーに分類する（表記は①～③の形【※必ずしも③まで必要としない入力内容や添付資料により臨機応変に】）。
・各ニーズは20〜100字程度で、児童・生徒の実態を柔らかく教育的にまとめる。
・具体的な出力の形【出力フォーマット】
対象児童（生徒）は現在、以下の状況である。【←必ずこの文言で始める】
① （入力内容を整理してまとめた課題、困難さ、特徴）
② （入力内容を整理してまとめた課題、困難さ、特徴）
③ （入力内容を整理してまとめた課題、困難さ、特徴）
※ 内容は「日常生活・学習・コミュニケーション・行動・身体・感覚・心理」などから、入力内容に応じて適切に選んでまとめること。

従って、以下の支援が必要である。【←必ずこの文言で始める】
① （①の状況に対応した支援目標や方法）
② （②の状況に対応した支援目標や方法）
③ （③の状況に対応した支援目標や方法）

支援に当たっては、以下の配慮が必要である。【←必ずこの文言で始める】
① （①の状況に対応した配慮）
② （②の状況に対応した配慮）
③ （③の状況に対応した配慮）

2.合理的配慮の実施内容
・上記①のニーズに対応した合理的配慮を少なくとも2つ以上提案する。
・シンプルに2つ以上箇条書きで示す。（例：「具体物を示し、視覚的な支援を行う。」、「落ち着ける環境設定にする。」、「生徒が好きな感触の本や音の出る玩具等を教室内に用意し、心理的安定を図る。」など）。

【条件】：
- 添付資料がある場合（「プランA」や「個別の教育支援計画」）にある書き方を参考にしてください。
- この後、この「特別な教育的ニーズ」を「特別な教育的ニーズ」⇒「所属校の支援目標」及び「各目標に連動した支援内容」⇒「指導方針」（プランBや個別の指導計画）の順で具体化していくのでそのつもりで抽象的に表現する。
- 「～です。～ます。」調ではなく、「～である。」調で文章を作成してください。
- 各項目の文量は200〜300文字程度で、柔らかく教育的な表現で整えてください。
"""

            if is_excel_mode:
                # --- Excelモード: JSON化 + 整形ルール + 「タイトル不要」の指示 ---
                prompt_text = f"""以下の実態や課題をもとに、特別支援教育に関する「プランA」の項目を作成してください。
ただし、出力形式は指定の【JSON形式】のみとします。前置きや解説は不要です。

【入力】実態や課題：
{jittai_1}

{common_instructions}

【重要：出力形式と整形ルール】
上記の内容を作成した上で、以下のJSONフォーマットに格納して出力してください。
JSONのvalue（値）については、以下のルールを厳守してください。

1. **タイトルの除外**：
   Excelのセルに既にヘッダーがあるため、出力テキストには「1.特別な教育的ニーズ」や「2.合理的配慮の実施内容」といった**タイトル行を含めないでください**。中身の文章のみを記述してください。

2. **改行の厳守**：
   Excelのセル内で見やすくするため、以下の箇所の直前には**必ず改行**を入れてください。
   - 「①」「②」「③」などの番号の直前
   - 「・」などの箇条書き記号の直前
   - 「従って、」「支援に当たっては、」などの段落の変わり目の直前
   
3. **余計な記号の禁止**：
   - Gemini等で表示される**「**」（太字強調）などのMarkdown記号は一切使用しないでください**。プレーンテキストのみにしてください。

【JSON出力フォーマット】
{{
  "needs": "特別な教育的ニーズの内容のみ（タイトル不要、整形ルールに従う）",
  "accommodations": "合理的配慮の実施内容のみ（タイトル不要、整形ルールに従う）"
}}
"""
            else:
                # --- 通常モード ---
                prompt_text = f"""以下の実態や課題をもとに、特別支援教育に関する「プランA」の以下の項目を作成してください。

【入力】実態や課題：
{jittai_1}
{common_instructions}
"""

            st.code(prompt_text, language="text")
            st.success("👆 右上のアイコンでコピーし、AIに貼り付けてください。")

# ==========================================
# プロンプト②
# ==========================================
with st.expander("プロンプト②【プランA：所属校の支援】"):
    with st.container(border=True):
        st.subheader("プロンプト②【プランA：所属校の支援】")
        needs_2 = st.text_area("✅ プロンプト①でAIが生成した「特別な教育的ニーズ」を貼り付け", value="（例）対象児童は現在、感覚過敏があり...", height=150, key="needs_2")
        
        if st.button("プロンプト② を生成", key="btn_2", use_container_width=True):
            
            # ★ 共通の指示内容
            common_instructions_2 = """
【出力項目】：
・1所属校の支援目標・機関名、2支援内容のプロンプトを生成　※機関名はそういう表記がされているだけで、意味は無視して良いタイトルはそのままで。
1.所属校の支援目標・機関名  
（特別な教育的ニーズの①～③に対応して作成。②までしかない場合は②まででよい。）

①目標：（30字以内程度・短くてもよい）  
②目標：（30字以内程度・短くてもよい）    
③目標：（30字以内程度・短くてもよい）  

2.支援内容  
（上記の①～③の目標それぞれに対応して作成。）

①（50字以内程度で、学校現場で実践可能な支援内容を記載）  
②（50字以内程度で、学校現場で実践可能な支援内容を記載）  
③（50字以内程度で、学校現場で実践可能な支援内容を記載）

・具体的な出力の形【出力フォーマット】 
このような抽象的な表現でよい（次のプロンプトでより具体化するため）。  

【条件】：
- 「特別な教育的ニーズ」と対応が分かるように①～③の番号を揃えること。  
- 各文は短くてもよいが、教育的で柔らかい表現にすること。  
- 「～です。～ます。」調ではなく、「～である。」調で統一すること。  
- 添付資料（「プランA」や「個別の教育支援計画」）の書き方を参考にしてよい。  
- ここでは抽象的にまとめ、**次の段階（プランBなど）で具体化していくための基礎**として作成すること。
"""
            
            if is_excel_mode:
                # --- Excelモード: JSON化 + 整形指示 + 「タイトル不要」 ---
                prompt_text = f"""以下の「特別な教育的ニーズ」に基づいて、「所属校による支援計画（プランA）」の項目を作成してください。
ただし、出力形式は指定の【JSON形式】のみとします。前置きや解説は不要です。

【参考】特別な教育的ニーズ：
{needs_2}

{common_instructions_2}

【重要：出力形式と整形ルール】
上記の内容を作成した上で、以下のJSONフォーマットに格納して出力してください。
JSONのvalue（値）については、以下のルールを厳守してください。

1. **タイトルの除外**：
   「1.所属校の支援目標・機関名」や「2.支援内容」といった**タイトル行を含めないでください**。
   
2. **改行の厳守**：
   Excelのセル内で見やすくするため、「①」「②」「③」などの番号の直前には**必ず改行**を入れてください。
   
3. **余計な記号の禁止**：
   - **「**」（太字強調）などのMarkdown記号は一切使用しないでください**。

【JSON出力フォーマット】
{{
  "goals": "支援目標の内容のみ（タイトル不要、整形ルールに従う）",
  "support": "支援内容の内容のみ（タイトル不要、整形ルールに従う）"
}}
"""
            else:
                # --- 通常モード ---
                prompt_text = f"""以下の「特別な教育的ニーズ」に基づいて、「所属校による支援計画（プランA）」の項目を作成してください。

【参考】特別な教育的ニーズ：
{needs_2}
{common_instructions_2}
"""

            st.code(prompt_text, language="text")
            st.success("👆 右上のアイコンでコピーし、AIに貼り付けてください。")

# ==========================================
# プロンプト③
# ==========================================
with st.expander("プロンプト③【プランB：指導方針・7項目の実態】"):
    with st.container(border=True):
        st.subheader("プロンプト③【プランB：指導方針・7項目の実態】")
        jittai_3 = st.text_area("✅ お子さんの実態や課題を入力（プロンプト①と同じでOK）", value="視力が弱い、落ち着きがない、疲れやすい、音に敏感、話しかけられると混乱する、同じ行動を繰り返す", height=100, key="jittai_3")
        
        if st.button("プロンプト③ を生成", key="btn_3", use_container_width=True):
            
            # ★ 共通の指示内容（例示・条件を含む全て）
            common_instructions_3 = """
【出力項目】

1.指導方針
・「特別な教育的ニーズ①～③」と「所属校の支援目標①～③」を踏まえ、より具体的な指導の方向性を示す。
・児童生徒の実態を冒頭で丁寧に描写し、その後に必要な指導内容を①～③に整理する。
・全体を通して300〜500字程度とし、教育的で柔らかい表現を用いる。
・文章が長くなる場合は複数回に分けて出力してよい。

2.実態（特別支援における7区分〔箇条書きで、例：【着替え】・～の形〕）
・以下の区分ごとに項目をもうけて（内容によって項目の数に偏りがあってもよい）記述する。各区分に例を設けるので参考にしてください
① 健康の保持（日常生活面、健康面など）【着替え】【食事】【排泄】【健康上の配慮】
（この区分は多めに項目を設ける【例：【着替え】・ズボンにはワッペンをつけると前後がわかって履くことができる。・言葉かけを受けて脱いだ服の裏返しを直すことができる。・登校時はオムツを着用している。登校後に布パンツに履き替えている。・下校時は布パンツを使用する。デイサービスにより帰り際にトイレに行かせる。【食事】・食事は少量ずつ別皿に入れて提供している。また、左手をお椀に添えて、右手でスプーンや補助具付きの箸を使用して食べることができる。【排泄】・立ち便器を使うことができる。【健康上の配慮】・春は花粉症の薬を服用している。など】）
② 心理的な安定（情緒面、状況の理解など）【苦手な状況】
【例：【苦手な状況】・急な音や大きな音は苦手で、驚いて不安定になることがある。・気持ちが不安定な時は、首に掴みかかろうとしたり、噛もうとしたりすることがある。・暑い季節は苦手で、夏場は気持ちが不安定になることが多い。汗を拭くことにより気持ちが落ち着くことがある。・気持ちが不安定な時に、イヤーマフを着用している。・気持ちが不安定なときに深呼吸を促したり、教員が胸や背中をトントン叩いたりすると落ち着くことがある。（感覚統合、副交感神経）【その他】・歩いたり、身体を動かしたりすることが好きで、散歩をすることで気持ちを切り替えられることがある。・見通しの持ちやすい課題には３０分程度離席せずに取り組むことができる。・見通しの持ちにくい場面や、気持ちが向かない活動の時にはトイレに行こうとすることがある。・怒ると近くにある物を噛む癖がある。(かなり減ってきた)】
③ 人間関係の形成（人とのかかわり、集団への参加など）【大人や友達との関わり】【集団参加】
【例：【大人や友達との関わり】・教員からの呼びかけに反応し、行動することができる。・身近な大人の膝の上に乗ろうとしたり、抱き着こうとしたりするなどの身体・接触が好きである。【集団参加】・誰とでも手を繋いだり、関わったりすることができる。】
④ 環境の把握（感覚の活用、認知面、学習面など）【学習の様子】
【例：【学習の様子】・シールを自分で好きなように貼れる。丸い紙やタイルを一直線に貼れる。・色の区別ができる。・３０面までのパズルができる。・ひらがなやイラストのマッチングができる。】
⑤ 身体の動き（運動・動作、作業面など）【身体の動き】【手指の操作】
【例：【身体の動き】・音楽が好きで、聴きながら身体を動かすことができる。・ブランコに一人で乗れる。・ボールを投げることができる。支援を受けてボールを蹴ることができる。【手指の操作】・利き手がまだ定まっていないが、右手を日常的に使用する。・支援を受けてハサミやのりなどの道具を使うことができる。・粗大模倣は、教員の手本や映像を観たりしながら行うことができる。・つまむ、ひねる、回す、押すなど手指を使った活動ができる。・爪がないため微細な動きは苦手である。・自助箸を使用して食事が取れる。】
⑥ コミュニケーション（意思の伝達、言語の形成など）【ｺﾐｭﾆｹｰｼｮﾝの理解】【ｺﾐｭﾆｹｰｼｮﾝの表出】
【例：【ｺﾐｭﾆｹｰｼｮﾝの理解】・教員の指示をある程度理解していて、指示通り動くことができる。【ｺﾐｭﾆｹｰｼｮﾝの表出】・促されると､「ちょうだい」のサインを出すことができる。サインと一緒に・口を動かすことができる。サインを出している大人に物を渡すことができる。・教員を呼ぶときに、肩をトントンと叩くことができる。・「トイレ」「ごめん」など簡単な言葉の発語ができる。・排泄の意思表示は「トイレ」と伝えることができる。・口形模倣や単音の発声はできるようになってきている。】
⑦ その他（性格、行動特徴、興味関心など）【興味関心】
【例：【興味関心】・活動中に身体を大きく左右に動かしたり飛び跳ねたりするときがある。・おもちゃを床にたたきつけるように投げる。】
・【入力】の実態・課題をもとに、可能な限り多くの具体的内容を盛り込む。
・1回で書ききれない場合は、何回になってもよいので、複数回に分けて出力する。

【条件】：
- 添付資料があれば参考にしてください（プランBや個別の指導計画）。
- 「～です。～ます。」調ではなく、「～である。」調で文章を作成してください。
- 【入力】の内容だけでなく、これまでの特別な教育的ニーズと所属校の支援目標から発展させて作成する。
- 「1.指導方針」特別な教育的ニーズと所属校の支援目標より具体的な内容にする
- 「1.指導方針」は特別な教育的ニーズと所属校の支援目標の①～③に連動した形で（②までしかない場合は同様に①、②でOK）
- ① 指導方針の例：「現在、本生徒は、右目の視力が無く左目も視力が弱く視野が狭い。また、聴力は、全く聞こえていないと思われ、日常生活の殆どを教員に任せきりにしていたり、教員に対して常に触れ合い・揺さぶりなどの刺激を求めている。また、不適切な行動が多く見られ、刺激を求めて、自分の顔に唾や水・おしっこをかけたり、吹き出したりすることや物を投げたり、倒したり衝動的な行動、人とのかかわりの中で抓ったり、叩いたり、髪の毛を引っ張ったりすることがある。コミュケーション面では、手話によるサインが２～３個理解できる。また、自分で1個「トイレ」のサインを行う時がある。その他、相手の手を取り直接手を動かしての要求をすることが多い。
従って、以下の①～③の指導が必要だと考える。
　①トイレや食事、着替えなど日常生活の中で、教員に頼らずにひとりでできることを増やす。
　②不適切行動が起きた時に繰り返さないようにする。
　③絵カードや手話による要求を増やす。
これらの指導を元に本生徒の生活面での自立、行動面・コミュニケーション面での成長につなげていく。」
- 1回で書ききれない場合は、何回になってもよいので、複数回に分けて出力する。
- 指導方針は全体的な視点で、各実態は200〜300文字で丁寧に描写してください。
- **改行の厳守**：
   Excelのセル内で見やすくするため、「・」などの直前には**必ず改行**を入れてください。  
- **余計な記号の禁止**：
   - **「**」（太字強調）などのMarkdown記号は一切使用しないでください**
"""

            if is_excel_mode:
                # --- Excelモード: 内容維持 + JSON化 + 整形指示 + 「タイトル不要」 ---
                prompt_text = f"""以下の実態・課題をもとに、特別支援計画「プランB」の項目を作成してください。
ただし、出力形式は指定の【JSON形式】のみとします。前置きや解説は不要です。

【入力】実態・課題：
{jittai_3}

{common_instructions_3}

【重要：出力形式と整形ルール】
上記の詳細な指示に基づき内容を作成した上で、以下のJSONフォーマットに格納して出力してください。
JSONのvalue（値）となる文字列については、以下の整形ルールを厳守してください。

1. **タイトルの除外**：
   Excelのセルに既にヘッダーがあるため、「1.指導方針」や「① 健康の保持」といった**タイトルや項目名を含めないでください**。その項目の実態（内容）のみを記述してください。

2. **改行の厳守**：
   Excelのセル内で見やすくするため、「・」や「【項目名】」の直前には**必ず改行**を入れてください。
   
3. **余計な記号の禁止**：
   - **「**」（太字強調）などのMarkdown記号は一切使用しないでください**。

【JSON出力フォーマット】
{{
  "policy": "指導方針の内容のみ（タイトル不要、整形ルールに従う）",
  "status_1": "健康の保持の実態のみ（「① 健康の保持」は含めない、整形ルールに従う）",
  "status_2": "心理的な安定の実態のみ（「② 心理的な安定」は含めない）",
  "status_3": "人間関係の形成の実態のみ（「③ 人間関係の形成」は含めない）",
  "status_4": "環境の把握の実態のみ（「④ 環境の把握」は含めない）",
  "status_5": "身体の動きの実態のみ（「⑤ 身体の動き」は含めない）",
  "status_6": "コミュニケーションの実態のみ（「⑥ コミュニケーション」は含めない）",
  "status_7": "その他の実態のみ（「⑦ その他」は含めない）"
}}
"""
            else:
                # --- 通常モード ---
                prompt_text = f"""以下の実態・課題をもとに、特別支援計画「プランB」の項目を作成してください。

【入力】実態・課題：
{jittai_3}
{common_instructions_3}
"""
            st.code(prompt_text, language="text")
            st.success("👆 右上のアイコンでコピーし、AIに貼り付けてください。")

# ==========================================
# プロンプト④
# ==========================================
with st.expander("プロンプト④【個別の指導計画：目標と手立て】"):
    with st.container(border=True):
        st.subheader("プロンプト④【個別の指導計画：目標と手立て】")
        st.write("個別の指導計画の目標と具体的な手立てを、選択された教科ごとに生成します。")

        # 教科の選択
        subject_options = [
            "自立活動","日常生活の指導","職業","生活単元学習","作業学習","国語", "算数", "美術", "理科", "社会", "音楽", "図画工作", "体育", "家庭", 
            "外国語活動", "総合的な学習の時間", "自立活動", "日常生活の指導","保険","数学"
        ]
        selected_subjects = st.multiselect("✅ 目標と手立てを作成する教科を選択（複数選択可）", subject_options, key="selected_subjects_4")

        # 実態や課題の入力（各教科用）
        jittai_inputs_4 = {}
        st.markdown("---")
        st.subheader("💡 各教科の実態や課題を入力してください")
        for subject in selected_subjects:
            jittai_inputs_4[subject] = st.text_area(f"✅ 【{subject}】に関する内容や児童生徒の実態や課題を入力",
                                                      value="（例：ミニトマトの植生を行った（種上・観察・記録・収穫・調理）",
                                                      height=100, key=f"jittai_4_{subject}")

        # 指導方針の入力（参考情報として）
        shido_hoshin_4 = st.text_area("✅ 実態や課題",
                                      value="（例：文字を読むことに抵抗がある、数の概念が理解しづらい、落ち着いて座っていられない、友達とのコミュニケーションが苦手など）、不適切行動が起きた時に繰り返さないようにする。絵カードや手話による要求を増やす。など）",
                                      height=100, key="shido_hoshin_4_global")

        if st.button("プロンプト④ を生成", key="btn_4", use_container_width=True):
            full_prompt_output = []
            for subject in selected_subjects:
                num_items = 3 if subject in ["自立活動", "日常生活の指導"] else 2
                current_jittai = jittai_inputs_4.get(subject, "")

                prompt_for_subject = f"""
以下の情報をもとに、個別の指導計画における【{subject}】の目標と手立てを作成してください。

【入力】
教科：{subject}
教科の内容：{current_jittai}
参考指導方針：{shido_hoshin_4}

【出力項目】
1. 目標（{num_items}つ）：
   ・各目標は30字以内程度で、お子さんが達成すべき具体的な行動や状態を示すこと。
   ・教育的で柔らかい表現にすること。
   ・「～できる」調にする。
   ・教科が「自立活動」の場合は、【教育活動全般】と【時間における指導】の二つから目標をそれぞれ設定する（例：【教育活動全般】自分の意見や考えを相手に正確に伝える。【時間における指導】・友達と意思の疎通を図りながら言葉のやりとりをしたり、ゲームをしたりすることができる。・友だちと協力して活動に取り組むことができる。・器具を使ってバランスを取ったり姿勢を保持したりすることができる。）。
   ・各教科の例（美術：目標「・様々な素材や色から、好きなものを選び作品作りができる。・様々な道具を使って作品作りができる。・鑑賞を通して、自分の好きな作品を選ぶことができる。」）
   
2. 手立て（{num_items}つ）：
   ・教科が「自立活動」の場合は、【教育活動全般】と【時間における指導】の二つから手立てをそれぞれ目標に連動する形で設定する（例：【教育活動全般】・手話や文字カードや音声アプリの活用する。【時間における指導】・言葉でのやりとりなどで適切なコミュニケーションができた場合に称賛する。・色々な友達と活動できるようにする。・感覚刺激により協調運動の向上をはかる。）。
   ・各手立ては30字から50字程度で、目標達成のために学校現場で実践可能な具体的な支援内容や方法を示すこと。
   ・お子さんの実態や課題、指導方針を考慮し、個別具体的な内容にすること。
   ・文の最後は「～する。」「～を促す」など手立てに相応しい形にする。
   ・各教科の例（美術：手立て「・様々な素材や道具を用意することで、その中から選べるようにする。・使い方の掲示、教員と一緒につかったりする。・友達の作品を分かりやすく並べ、好きな作品を選べるようにする。」）
  

【出力フォーマット例】
【{subject}】
目標：
・（目標1：30字以内）
・（目標2：30字以内）
{"・（目標3：30字以内）" if num_items == 3 else ""}

手立て：
・（手立て1：30～50字）
・（手立て2：30～50字）
{"・（手立て3：30～50字）" if num_items == 3 else ""}

【条件】：
- 添付資料（個別の指導計画など）がある場合は、その書き方を参考にしてください。
- 他の教科の目標や手立ては出力せず、指定された【{subject}】のみを出力してください。
- 目標と手立ての数は、自立活動と日常生活の指導は3つ以上、それ以外は2つとしてください。
- 目標と手立ての内容は、入力された実態や課題、参考指導方針と連動させて具体的に記述してください。
- **改行の厳守**：
   Excelのセル内で見やすくするため、「・」などの直前には**必ず改行**を入れてください。  
- **余計な記号の禁止**：
   - **「**」（太字強調）などのMarkdown記号は一切使用しないでください**
"""
                full_prompt_output.append(prompt_for_subject)

            st.subheader("📄 生成されたプロンプト④（コピーして使ってください）")
            if not full_prompt_output:
                st.warning("教科が選択されていません。")
            else:
                st.code("\n---\n".join(full_prompt_output), language="text") # 各教科のプロンプトを区切り線で結合して表示

# ==========================================
# プロンプト⑤
# ==========================================
with st.expander("プロンプト⑤【個別の指導計画：評価】"):
    with st.container(border=True):
        st.subheader("プロンプト⑤【個別の指導計画：評価】")
        st.write("指導計画を基に、活動の様子を評価する文章を作成します。")

        use_file_4 = st.checkbox(
            "AIにWord/Excel等のファイルを添付して、主たる情報源として利用する", 
            key="use_file_4"
        )
        st.caption("💡 ファイルを添付する場合、下のテキストエリアは補足として利用できます。")

        reference_text_4 = st.text_area(
            "✅ 参考にする指導計画のテキスト（補足など）",
            value="（例：個別の指導計画の「指導の目標および内容」の全文や、特に見てほしい部分など）",
            height=150,
            key="reference_4"
        )
        
        evaluation_activities_4 = st.text_area(
            "✅ 教科ごとの「できたこと・活動の様子」を具体的に入力",
            value="【自立活動】：・教員の誘導で肩や首の力を抜き、胸を張った姿勢で活動できた。・片手で支えながら片足立ちができた。\n【国語】：自分の名前を丁寧になぞり書きできた。",
            height=150,
            key="evaluation_activities_4"
        )

        if st.button("プロンプト⑤ を生成", key="btn_5", use_container_width=True):
            prompt_intro_4 = ""
            prompt_main_source_4 = ""

            if use_file_4:
                prompt_intro_4 = "添付した指導計画のファイルを主たる情報源とし、"
                if reference_text_4.strip() and reference_text_4 != "（例：個別の指導計画の「指導の目標および内容」の全文や、特に見てほしい部分など）":
                    prompt_main_source_4 = f"以下の【参考テキスト】も補足情報として考慮した上で、"
                else:
                      prompt_main_source_4 = "以下の"

            else:
                prompt_intro_4 = "以下の【指導計画のテキスト】を主たる情報源として、"
            
            prompt_text_part_4 = f"""【指導計画のテキスト】：
{reference_text_4}
""" if not use_file_4 or (use_file_4 and reference_text_4.strip() and reference_text_4 != "（例：個別の指導計画の「指導の目標および内容」の全文や、特に見てほしい部分など）") else ""


            prompt_full_4 = f"""{prompt_intro_4}{prompt_main_source_4}【できたこと・活動の様子】とを関連付けながら、教科ごとの「個別の指導の評価」（振り返り文）を作成してください。
{prompt_text_part_4}
【できたこと・活動の様子】：
{evaluation_activities_4}

【出力ルール】：
- まず、主たる情報源（添付ファイルまたは上記テキスト）を読み込み、そこに書かれている目標や内容を完全に理解してください。
- 「～です。～ます。」調ではなく、「～である。」調で文章を作成してください。
- 文の最後は「～できた。」で終わらせる。
- その上で、【できたこと・活動の様子】が、計画のどの目標・内容に対応するのかを分析し、目標の達成度合いが分かるように評価文を作成してください。
- 計画で言及されているすべての教科・領域について、評価文を個別に出力してください。
- 各教科について、【教科名の見出し】と200～300文字程度の評価文を作成してください。
- 文体は、実務で使用されるような柔らかく教育的な表現にしてください。
- 各教科の例（美術：「・仙台七夕祭りの吹流し作りでは、折り染めに取り組んだ。染める色を３つ選択し、染める手元をよく見て色の滲みに注目して染めることができた。・土器作りではへらや縄、貝殻やビー玉などを粘土に押し付けて模様をつけることができた・土器の鑑賞では、友達の作品の中から気にいったものを２つ選ぶことができた。」） """           

            st.subheader("📄 生成されたプロンプト⑤（コピーして使ってください）")
            st.code(prompt_full_4, language="text")


# ==========================================
# プロンプト⑥
# ==========================================
with st.expander("プロンプト⑥【前期・後期の所見】"):
    with st.container(border=True):
        st.subheader("プロンプト⑥【前期・後期の所見】")
        st.write("評価文や計画書を基に、総合的な所見を作成します。")

        term_choice = st.radio("どちらの所見を作成しますか？", ("前期", "後期／学年末"), key="term_choice", horizontal=True)

        use_file_5 = st.checkbox(
            "AIに評価文等のファイルを添付して、主たる情報源として利用する",
            key="use_file_5"
        )
        st.caption("💡 ファイルを添付する場合、下のテキストエリアは補足として利用できます。")

        reference_text_5 = st.text_area(
            "✅ 参考にする評価文や計画のテキスト（補足など）",
            value="（例：プロンプト④で作成した評価文の全体、または特に見てほしい部分など）",
            height=200,
            key="reference_5"
        )
        
        shoken_input = st.text_area(
            "✅ 所見で特に強調したいポイントを入力（箇条書きでOK）",
            value="- 大きなけがもなく元気に登校できたことの喜び。\n- 友人との関わりが前向きになった点。\n- 宿泊学習などの大きな行事を乗り越えた自信。\n- 家庭との連携への感謝。",
            height=150,
            key="shoken_input"
        )

        if st.button("プロンプト⑥ を生成", key="btn_6", use_container_width=True):
            prompt_intro_5 = ""
            prompt_main_source_5 = ""

            if use_file_5:
                prompt_intro_5 = "添付したファイル（評価文や計画書など）を主たる情報源とし、"
                if reference_text_5.strip() and reference_text_5 != "（例：プロンプト④で作成した評価文の全体、または特に見てほしい部分など）":
                    prompt_main_source_5 = "以下の【参考テキスト】も補足情報として考慮し、"
                else:
                    prompt_main_source_5 = ""
            else:
                prompt_intro_5 = "以下の【参考テキスト】を主たる情報源として、"

            prompt_text_part_5 = f"""【参考テキスト】：
{reference_text_5}
""" if not use_file_5 or (use_file_5 and reference_text_5.strip() and reference_text_5 != "（例：プロンプト④で作成した評価文の全体、または特に見てほしい部分など）") else ""

            if term_choice == "前期":
                specific_conditions = """- 「前期は、〜」といった書き出しで始めてください。
- 200～400文字程度の文章量で作成してください。
- 文末には「後期も引き続き、ご家庭と連携を取りながら、成長を見守っていきたいと思います。よろしくお願いいたします。」など、後期への連携を意識した一文を入れてください。"""
            else: # 後期／学年末
                specific_conditions = """- 「この1年間で〜」や「いよいよ来年度は〜」など、年度の区切りを感じさせる書き出しにしてください。
- 200〜450文字程度の文章量で作成してください。
- 文末には、保護者への感謝と、次年度に向けた応援の言葉を必ず含めてください。"""

            prompt_full_5 = f"""{prompt_intro_5}{prompt_main_source_5}さらに【強調したいポイント】を盛り込みながら、保護者向けの「{term_choice}の所見」を作成してください。
{prompt_text_part_5}
【強調したいポイント】：
{shoken_input}

【全体の共通条件】：
- 主たる情報源（添付ファイルまたは上記テキスト）から全体的な成長の様子を読み取り、【強調したいポイント】を特に意識して、自然な文章を作成してください。
- 添付資料内に具体的事象（例：大きな行事の達成、友人とのやりとり、けがの有無）がある場合は具体例を1つ以上挙げて所見に反映すること。
- 丁寧な語り口で、前向きな表現を心がけてください。
- 保護者の方が読んで、お子さんの具体的な成長が伝わるような文章にしてください。
- 「～です。～ます。」調ではなく、「～である。」調で文章を作成してください。
- もし【入力】や添付資料の内容が極端に短い・情報不足な場合は、前期は200字、後期は200字を下回らないように、実務的に妥当な記述を参考テキストに基づき推測して補完すること

【{term_choice}用の個別条件】：
{specific_conditions}"""

            st.subheader(f"📄 生成されたプロンプト⑥（{term_choice}の所見用）")
            st.code(prompt_full_5, language="text")

st.markdown("---")

# ============================================================
# ★★★ 岩槻はるかぜ特別支援学校の人だけ機能 (Excelモード時のみ表示) ★★★
# ============================================================
if is_excel_mode:
    st.header("✨ 岩槻はるかぜ特別支援学校の人だけ機能")
    st.info("この機能は「Excel作成モード」選択時のみ表示されます。\nAIが出力したJSONコードを以下の欄に貼り付けるだけで、Excelに自動入力されます。")
    st.warning("※WEBアプリと同じフォルダに「プラン.xlsx」が必要です。")

    c_iwatsuki_1, c_iwatsuki_2 = st.columns(2)
    
    # --- 左カラム：プロンプト①と② (プランA) ---
    with c_iwatsuki_1:
        st.markdown("#### 📄 プロンプト①の結果 (プランA)")
        json_input_1 = st.text_area("JSON貼り付け (needs, accommodations)", height=200, placeholder='{\n  "needs": "...",\n  "accommodations": "..."\n}')

        st.markdown("#### 📄 プロンプト②の結果 (プランA)")
        json_input_2 = st.text_area("JSON貼り付け (goals, support)", height=200, placeholder='{\n  "goals": "...",\n  "support": "..."\n}')

    # --- 右カラム：プロンプト③ (プランB) ---
    with c_iwatsuki_2:
        st.markdown("#### 📄 プロンプト③の結果 (プランB)")
        json_input_3 = st.text_area("JSON貼り付け (policy, status_1...)", height=500, placeholder='{\n  "policy": "...",\n  "status_1": "...",\n  ...\n}')

    # --- Excel生成ボタン ---
    if st.button("🚀 Excelに書き出してダウンロード", type="primary", use_container_width=True):
        try:
            template_path = "プラン.xlsx"
            if not os.path.exists(template_path):
                st.error(f"❌ エラー: '{template_path}' が見つかりません。アプリと同じ場所に配置してください。")
            else:
                wb = openpyxl.load_workbook(template_path)
                
                # -------------------------------------
                # A. プロンプト①の書き込み (プランＡ)
                # -------------------------------------
                if json_input_1.strip():
                    try:
                        clean_json_1 = re.sub(r"```json\s*|\s*```", "", json_input_1).strip()
                        s_idx = clean_json_1.find('{')
                        e_idx = clean_json_1.rfind('}') + 1
                        if s_idx != -1 and e_idx != -1:
                            clean_json_1 = clean_json_1[s_idx:e_idx]
                        data_1 = json.loads(clean_json_1)
                        
                        if "プランＡ" in wb.sheetnames:
                            ws_a = wb["プランＡ"]
                            safe_write(ws_a, 'D12', data_1.get('needs', ''))
                            safe_write(ws_a, 'D15', data_1.get('accommodations', ''))
                        else:
                            st.warning("⚠️ シート「プランＡ」が見つかりません。")
                    except json.JSONDecodeError:
                        st.error("❌ プロンプト①のJSON解析エラー。貼り付け内容を確認してください。")

                # -------------------------------------
                # B. プロンプト②の書き込み (プランＡ)
                # -------------------------------------
                if json_input_2.strip():
                    try:
                        clean_json_2 = re.sub(r"```json\s*|\s*```", "", json_input_2).strip()
                        s_idx = clean_json_2.find('{')
                        e_idx = clean_json_2.rfind('}') + 1
                        if s_idx != -1 and e_idx != -1:
                            clean_json_2 = clean_json_2[s_idx:e_idx]
                        data_2 = json.loads(clean_json_2)
                        
                        if "プランＡ" in wb.sheetnames:
                            ws_a = wb["プランＡ"]
                            safe_write(ws_a, 'D18', data_2.get('goals', ''))
                            safe_write(ws_a, 'E18', data_2.get('support', ''))
                    except json.JSONDecodeError:
                        st.error("❌ プロンプト②のJSON解析エラー。貼り付け内容を確認してください。")

                # -------------------------------------
                # C. プロンプト③の書き込み (プランＢ(実態))
                # -------------------------------------
                if json_input_3.strip():
                    try:
                        clean_json_3 = re.sub(r"```json\s*|\s*```", "", json_input_3).strip()
                        s_idx = clean_json_3.find('{')
                        e_idx = clean_json_3.rfind('}') + 1
                        if s_idx != -1 and e_idx != -1:
                            clean_json_3 = clean_json_3[s_idx:e_idx]
                        data_3 = json.loads(clean_json_3)
                        
                        if "プランＢ(実態)" in wb.sheetnames:
                            ws_b = wb["プランＢ(実態)"]
                            safe_write(ws_b, 'C5', data_3.get('policy', ''))
                            safe_write(ws_b, 'D8', data_3.get('status_1', ''))
                            safe_write(ws_b, 'D10', data_3.get('status_2', ''))
                            safe_write(ws_b, 'D12', data_3.get('status_3', ''))
                            safe_write(ws_b, 'D14', data_3.get('status_4', ''))
                            safe_write(ws_b, 'D16', data_3.get('status_5', ''))
                            safe_write(ws_b, 'D18', data_3.get('status_6', ''))
                            safe_write(ws_b, 'D20', data_3.get('status_7', ''))
                        else:
                            st.warning("⚠️ シート「プランＢ(実態)」が見つかりません。")
                    except json.JSONDecodeError:
                        st.error("❌ プロンプト③のJSON解析エラー。貼り付け内容を確認してください。")

                # 保存とダウンロード
                excel_file = io.BytesIO()
                wb.save(excel_file)
                excel_file.seek(0)

                st.balloons()
                st.success("✨ Excelファイルの作成が完了しました！")
                st.download_button(
                    label="📥 更新された「プラン.xlsx」をダウンロード",
                    data=excel_file,
                    file_name="プラン_更新版.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )

        except Exception as e:
            st.error(f"予期せぬエラーが発生しました: {e}")

st.warning("""
**【利用上の注意】**
AIが生成する内容は、入力された情報に基づく提案であり、必ずしも正確性や適切性を保証するものではありません。自分の判断と合わせてご活用ください。
""")