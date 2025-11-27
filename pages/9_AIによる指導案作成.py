import streamlit as st
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.cell.cell import MergedCell
import json
import io
import os
import re
import base64
from pathlib import Path

# ==========================================
# 0. ページ設定
# ==========================================
st.set_page_config(
    page_title="MieeL - 指導案作成",
    page_icon="📝",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ==========================================
# 1. 画像処理 (ロゴ読み込み)
# ==========================================
def get_img_as_base64(file):
    try:
        # 画像パスを絶対パスで解決
        script_path = Path(__file__)
        app_root = script_path.parent.parent
        img_path = app_root / file
        
        with open(img_path, "rb") as f:
            data = f.read()
        return base64.b64encode(data).decode()
    except:
        return None

# ロゴファイル設定 (参考コードに合わせて配置)
logo_path = "MieeL2.png" 
logo_b64 = get_img_as_base64(logo_path)
logo_html = f'<img src="data:image/png;base64,{logo_b64}" class="logo-img">' if logo_b64 else '<div class="logo-placeholder">🌟</div>'


# ==========================================
# 2. デザイン定義 (MieeLデザイン + ぬるっと動くアニメーション)
# ==========================================
def load_css():
    st.markdown("""
        <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@400;700;900&display=swap" rel="stylesheet">
    """, unsafe_allow_html=True)
    
    css = f"""
    <style>
        /* --- 全体フォント --- */
        html, body, [class*="css"] {{
            font-family: 'Noto Sans JP', sans-serif !important;
            color: #333333 !important;
        }}

        /* --- 背景 (白92%透過・画像あり) --- */
        [data-testid="stAppViewContainer"] {{
            background-color: #ffffff;
            background-image: linear-gradient(rgba(255,255,255,0.92), rgba(255,255,255,0.92)), url("https://i.imgur.com/AbUxfxP.png");
            background-size: cover;
            background-attachment: fixed;
        }}

        /* --- 文字色 (濃紺・くっきり) --- */
        h1, h2, h3, h4, h5, h6 {{
            color: #0f172a !important; /* 濃いネイビーブラック */
            text-shadow: none !important;
        }}
        p, span, div, label {{
            color: #333333 !important;
            text-shadow: none !important;
        }}

        /* --- サイドバー (すりガラス効果) --- */
        [data-testid="stSidebar"] {{
            background-color: rgba(255, 255, 255, 0.85) !important;
            backdrop-filter: blur(20px);
            -webkit-backdrop-filter: blur(20px);
            border-right: 1px solid #e2e8f0 !important;
        }}
        [data-testid="stSidebar"] * {{
            color: #333333 !important;
        }}

        /* 
           ================================================================
           ★ アニメーション定義 (下からフワッと)
           ================================================================
        */
        @keyframes fadeInUp {{
            from {{ opacity: 0; transform: translateY(40px); }}
            to {{ opacity: 1; transform: translateY(0); }}
        }}

        /* 
           ================================================================
           ★ コンテナデザイン (白背景・影付き・アニメーション)
           ================================================================
           st.container(border=True) のスタイルをオーバーライド
        */
        [data-testid="stBorderContainer"] {{
            background-color: #ffffff !important;
            border: 2px solid #e2e8f0 !important; /* 薄いグレーの枠線 */
            border-radius: 15px !important;
            padding: 25px !important;
            margin-bottom: 20px !important;
            box-shadow: 0 4px 12px rgba(0,0,0,0.05) !important;
            
            /* アニメーション適用 */
            opacity: 0; 
            animation-name: fadeInUp;
            animation-duration: 0.8s;
            animation-fill-mode: forwards;
            animation-timing-function: cubic-bezier(0.2, 0.8, 0.2, 1);
        }}
        
        /* コンテナの出現タイミングをずらす */
        div.element-container:nth-of-type(1) [data-testid="stBorderContainer"] {{ animation-delay: 0.1s; }}
        div.element-container:nth-of-type(2) [data-testid="stBorderContainer"] {{ animation-delay: 0.3s; }}
        div.element-container:nth-of-type(3) [data-testid="stBorderContainer"] {{ animation-delay: 0.5s; }}

        [data-testid="stBorderContainer"]:hover {{
            border-color: #4a90e2 !important;
            background-color: #f8fafc !important;
            transform: translateY(-2px);
            box-shadow: 0 10px 25px rgba(74, 144, 226, 0.15) !important;
            transition: all 0.3s ease;
        }}

        /* --- ボタン --- */
        .stButton > button {{
            width: 100%;
            background-color: #ffffff !important;
            border: 2px solid #e2e8f0 !important;
            color: #4a90e2 !important;
            font-weight: bold !important;
            border-radius: 30px !important; /* 丸みを強く */
            padding: 10px !important;
            transition: all 0.3s ease !important;
        }}
        .stButton > button:hover {{
            background-color: #4a90e2 !important;
            color: #ffffff !important;
            border-color: #4a90e2 !important;
        }}
        
        /* Primaryボタン (強調) */
        .stButton > button[kind="primary"] {{
            background-color: #4a90e2 !important;
            color: #ffffff !important;
            border: 2px solid #4a90e2 !important;
        }}
        .stButton > button[kind="primary"]:hover {{
            background-color: #2563eb !important;
            border-color: #2563eb !important;
        }}

        /* --- 入力フォーム --- */
        .stTextInput input, .stTextArea textarea {{
            background-color: #ffffff !important;
            color: #333333 !important;
            border: 1px solid #cbd5e1 !important;
            border-radius: 8px !important;
        }}
        .stTextInput input:focus, .stTextArea textarea:focus {{
            border-color: #4a90e2 !important;
            box-shadow: 0 0 0 2px rgba(74,144,226,0.2) !important;
        }}

        /* --- ステップヘッダー --- */
        .step-header {{
            color: #0f172a !important;
            border-left: 5px solid #4a90e2;
            padding-left: 15px;
            margin-top: 40px;
            margin-bottom: 20px;
            font-weight: 900;
            font-size: 1.5em;
        }}

        /* --- 戻るボタン (リクエスト指定) --- */
        .back-link {{
            margin-bottom: 20px;
        }}
        .back-link a {{
            display: inline-block;
            padding: 10px 20px;
            background: #ffffff;
            border: 2px solid #e2e8f0;
            border-radius: 25px;
            color: #4a90e2 !important;
            text-decoration: none;
            font-weight: bold;
            transition: all 0.3s;
            box-shadow: 0 2px 5px rgba(0,0,0,0.05);
        }}
        .back-link a:hover {{
            background: #4a90e2;
            color: #ffffff !important;
            border-color: #4a90e2;
            box-shadow: 0 4px 10px rgba(74, 144, 226, 0.2);
        }}

        /* --- ヘッダーレイアウト --- */
        .header-container {{
            display: flex;
            align-items: center;
            justify-content: center; /* ここを追加 */
            gap: 20px;
            margin-bottom: 30px;
            padding-bottom: 20px;
            border-bottom: 2px solid #f1f5f9;
            animation: fadeInUp 0.8s ease-out forwards;
        }}
        .logo-img {{
            width: 80px;
            height: auto;
            filter: drop-shadow(0 4px 6px rgba(0,0,0,0.1));
        }}
        .page-title {{
            font-size: 2.2rem;
            font-weight: 900;
            color: #0f172a;
            margin: 0;
            line-height: 1.2;
        }}
        
        /* コードブロック */
        code {{
            background-color: #f1f5f9 !important;
            color: #0f172a !important;
            border: 1px solid #e2e8f0;
            border-radius: 8px;
        }}
    </style>
    """
    st.markdown(css, unsafe_allow_html=True)

load_css()

# ==========================================
# 3. ユーティリティ関数
# ==========================================
def safe_write(ws, cell_address, value):
    """結合セルエラーを回避して書き込む関数"""
    try:
        if value is None:
            value = ""
        value = str(value)

        if isinstance(ws[cell_address], MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell_address in merged_range:
                    top_left_coord = merged_range.start_cell.coordinate
                    ws[top_left_coord] = value
                    ws[top_left_coord].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                    return
        else:
            ws[cell_address] = value
            ws[cell_address].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

    except Exception as e:
        st.warning(f"⚠️ セル {cell_address} への書き込み中に警告: {e}")

def generate_prompt_text(data):
    prompt = f"""
あなたは特別支援学校および公立学校における【熟練の教員】です。
以下の【授業情報】を基に、学習指導案に必要な情報を補完し、指定の【JSON形式】のみで出力してください。
前置きや解説は一切不要です。JSONデータだけを返してください。

■ 【授業情報】
[必須項目]
・学部学年: {data['grade']}
・教科単元: {data['subject']}
・日時: {data['date']}
・時間: {data['time']}
・場所: {data['place']}
・本時の内容: {data['content']}

[任意項目]
・目標: {data['goals_in'] if data['goals_in'] else "未定（文脈に合わせて最大3つ生成せよ）"}
・評価の基準: {data['eval_in'] if data['eval_in'] else "未定（3観点：知識・技能、思考判断表現、主体的態度を含めて生成せよ）"}
・学習内容のメモ: {data['flow_in'] if data['flow_in'] else "未定（自然な流れで構成せよ）"}
・備考: {data['remarks_in'] if data['remarks_in'] else "なし"}

■ 【出力フォーマット（厳守）】
以下のJSON構造を絶対に崩さずに返してください。
{{
  "basic_info": {{
    "grade": "{data['grade']}",
    "subject": "{data['subject']}",
    "date": "{data['date']}",
    "time": "{data['time']}",
    "place": "{data['place']}",
    "content": "{data['content']}"
  }},
  "goals": ["目標1", "目標2", "目標3"],
  "evaluation": ["評価基準1（知識技能）", "評価基準2（思考判断）", "評価基準3（主体性）"],
  "flow": [
    {{
      "time": "5",
      "activity": "導入：挨拶...",
      "notes": "配慮事項..."
    }},
    {{
      "time": "10",
      "activity": "展開1：...",
      "notes": "..."
    }}
  ],
  "materials": "準備物リスト",
  "remarks": "備考の内容（特になければ空欄でも可）"
}}
"""
    return prompt

def create_excel(template_path, json_data):
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
    except Exception as e:
        return None, f"テンプレート読み込みエラー: {e}"

    # データ書き込み
    bi = json_data.get('basic_info', {})
    safe_write(ws, 'C2', bi.get('grade', ''))
    safe_write(ws, 'I2', bi.get('subject', ''))
    safe_write(ws, 'C3', bi.get('date', ''))
    safe_write(ws, 'K3', bi.get('time', ''))
    safe_write(ws, 'N3', bi.get('place', ''))
    safe_write(ws, 'C4', bi.get('content', ''))

    goals = json_data.get('goals', [])
    for i in range(min(len(goals), 3)):
        safe_write(ws, f'C{5+i}', f"・{goals[i]}")

    evals = json_data.get('evaluation', [])
    for i in range(min(len(evals), 3)):
        safe_write(ws, f'C{8+i}', f"・{evals[i]}")

    flow_list = json_data.get('flow', [])
    current_row = 13
    for item in flow_list:
        safe_write(ws, f'A{current_row}', item.get('time', ''))
        safe_write(ws, f'B{current_row}', item.get('activity', ''))
        safe_write(ws, f'K{current_row}', item.get('notes', ''))
        current_row += 2

    safe_write(ws, 'N13', json_data.get('materials', ''))
    safe_write(ws, 'B33', json_data.get('remarks', ''))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, None

# ==========================================
# 4. メイン画面 UI
# ==========================================

# --- 戻るボタン (★正しいリンクに変更済み) ---
st.markdown('<div class="back-link"><a href="https://aspecialeducationapp-6iuvpdfjbflp4wyvykmzey.streamlit.app/" target="_self">« TOPページに戻る</a></div>', unsafe_allow_html=True)

# --- ヘッダーエリア ---
st.markdown(f"""
    <div class="header-container">
        {logo_html}
        <div>
            <h1 class="page-title">AI指導案作成エージェント</h1>
            <p style="color:#475569; margin:0; font-weight:bold;">プロンプト生成 ➡ AIに入力 ➡ Excel出力 の3ステップ</p>
        </div>
    </div>
""", unsafe_allow_html=True)

# --- AIリンクボタン ---
with st.container(border=True):
    st.markdown("### 🔗 まずはAIを開く")
    col_btn1, col_btn2 = st.columns(2)
    with col_btn1:
        st.link_button("🤖 ChatGPT を開く ↗", "https://chat.openai.com/", type="primary", use_container_width=True)
    with col_btn2:
        st.link_button("✨ Gemini を開く ↗", "https://gemini.google.com/", type="primary", use_container_width=True)

# --- Step 1: 情報入力 ---
st.markdown("<h3 class='step-header'>Step 1. 基本情報を入力</h3>", unsafe_allow_html=True)

with st.container(border=True):
    c1, c2, c3 = st.columns(3)
    with c1:
        in_grade = st.text_input("🎓 学部学年", "小学部 5年")
        in_date = st.text_input("📅 日時", "令和6年11月20日")
    with c2:
        in_subject = st.text_input("📚 教科単元", "生活単元学習「お祭りを開こう」")
        in_place = st.text_input("🏫 場所", "5年1組教室")
    with c3:
        in_time = st.text_input("⏰ 時間", "45分")
        in_content = st.text_input("📝 本時の内容", "模擬店の商品作り")

    st.markdown("---")
    # 詳細設定
    with st.expander("⚙️ 詳細設定（目標・評価・備考など） ※空欄でもAIが補完します", expanded=False):
        col_ex1, col_ex2 = st.columns(2)
        with col_ex1:
            in_goals = st.text_area("🎯 目標（最大3つ）", height=100, placeholder="例：\n・道具を正しく使うことができる\n・友達と協力することができる")
            in_eval = st.text_area("📊 評価の基準", height=100, placeholder="知識・技能、思考・判断・表現、主体的に取り組む態度の観点で生成されます。")
        with col_ex2:
            in_flow = st.text_area("💡 学習内容のメモ・ヒント", height=100, placeholder="授業の流れや、必ず入れたい活動があれば箇条書きで。")
            in_remarks = st.text_area("📌 備考（特記事項）", height=100, placeholder="Excelの下部（B33）に入力されます。")

# データをまとめる
input_data = {
    "grade": in_grade, "subject": in_subject, "date": in_date,
    "time": in_time, "place": in_place, "content": in_content,
    "goals_in": in_goals, "eval_in": in_eval, "flow_in": in_flow,
    "remarks_in": in_remarks
}

# --- Step 2: プロンプト生成 ---
st.markdown("<h3 class='step-header'>Step 2. プロンプトをコピー</h3>", unsafe_allow_html=True)

with st.container(border=True):
    if st.button("📋 プロンプトを作成する", type="primary", use_container_width=True):
        prompt_text = generate_prompt_text(input_data)
        st.code(prompt_text, language="text")
        st.success("👆 右上のアイコンでコピーし、ChatGPTやGeminiに貼り付けてください。")
    else:
        st.info("上のボタンを押すと、AIへの指令文が表示されます。")

# --- Step 3: AI出力貼り付け & Excel生成 ---
st.markdown("<h3 class='step-header'>Step 3. AIの回答を貼り付けてExcel作成</h3>", unsafe_allow_html=True)

with st.container(border=True):
    json_input_str = st.text_area("ここにAIからの回答（JSONコード）を貼り付け", height=300, placeholder='{\n  "basic_info": { ... },\n  "goals": [ ... ]\n}')

    if st.button("🚀 指導案Excelを出力する", type="primary", use_container_width=True):
        if not json_input_str.strip():
            st.error("⚠️ AIの回答が貼り付けられていません。")
        else:
            try:
                # JSONクリーニング
                clean_json = re.sub(r"```json\s*|\s*```", "", json_input_str).strip()
                start_idx = clean_json.find('{')
                end_idx = clean_json.rfind('}') + 1
                if start_idx != -1 and end_idx != -1:
                    clean_json = clean_json[start_idx:end_idx]
                
                data_dict = json.loads(clean_json)
                
                # テンプレート検索
                current_dir = os.path.dirname(os.path.abspath(__file__))
                base_dir = os.path.dirname(current_dir)
                template_file = os.path.join(base_dir, "指導案.xlsx")
                
                if not os.path.exists(template_file):
                    template_file = os.path.join(current_dir, "指導案.xlsx")

                if not os.path.exists(template_file):
                    st.error(f"❌ エラー: テンプレートファイルが見つかりません。\n{base_dir} または {current_dir} に '指導案.xlsx' を配置してください。")
                else:
                    # Excel生成
                    excel_data, err = create_excel(template_file, data_dict)
                    if err:
                        st.error(err)
                    else:
                        st.balloons()
                        st.success("✨ 指導案Excelが完成しました！")
                        st.download_button(
                            label="📥 完成した指導案をダウンロード",
                            data=excel_data,
                            file_name="完成_指導案.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            type="primary"
                        )
                        
            except json.JSONDecodeError:
                st.error("❌ JSON解析エラー: 貼り付けたテキストが正しいJSON形式か確認してください。")
            except Exception as e:
                st.error(f"❌ 予期せぬエラー: {e}")